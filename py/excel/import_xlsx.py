# -*- coding: utf-8 -*-

"""
upload and save
show in grid and editable
edit and save

json or sql?

"""

import openpyxl


def main():

    filename = "weekly_plan_template.xlsx"

    xls_book = openpyxl.load_workbook(filename=filename)

    sheet_names = xls_book.get_sheet_names()

    xls_sheet = xls_book.get_sheet_by_name(sheet_names[0])

    max_row = xls_sheet.max_row

    for rownum in range(1, max_row):
        # print rownum
        for colnum in range(9):
            val = xls_sheet.cell(row=rownum + 1, column=colnum + 1).value
            if val is None:
                break

            val_type = type(val)

            if isinstance(val, unicode):
                print(val.encode("utf8")),
            else:
                print val,
        print("\n")


if __name__ == "__main__":
    main()